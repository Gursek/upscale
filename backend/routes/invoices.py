from flask import Blueprint, request, jsonify, send_file, Response
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime, date, timedelta
from decimal import Decimal
from io import BytesIO
import json
from html import escape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.db import db
from models.product import Product
from models.invoice import Invoice, InvoiceItem
from models.user import User
from models.ejournal import EJournalEntry
from models.ejournal import ZReading
from services.tax import resolve_tax_classification, compute_invoice_totals
from services.business_time import (
    business_day_utc_bounds,
    business_today,
    to_business_iso,
    utc_now_naive,
)
from services.compliance import add_audit_event, canonical_json, seller_snapshot

invoices_bp = Blueprint("invoices", __name__)


def _invoice_to_dict(inv: Invoice) -> dict:
    seller = json.loads(inv.seller_snapshot_json or "{}")
    reset_counter = int(seller.get("reset_counter") or 0)
    return {
        "id": inv.id,
        "invoice_number": inv.invoice_number,
        "invoice_type": inv.invoice_type,
        "date_time": to_business_iso(inv.date_time),
        "cashier_name": inv.cashier_name,
        "payment_mode": inv.payment_mode,
        "cash_tendered": float(inv.cash_tendered or 0),
        "change_amount": float(inv.change_amount or 0),
        "buyer_name": inv.buyer_name,
        "buyer_address": inv.buyer_address,
        "buyer_tin": inv.buyer_tin,
        "buyer_business_style": inv.buyer_business_style,
        "seller": seller,
        "display_invoice_number": f"{reset_counter:02d}-{inv.invoice_number}",
        "vatable_sales": float(inv.vatable_sales),
        "vat_amount": float(inv.vat_amount),
        "vat_exempt_sales": float(inv.vat_exempt_sales),
        "zero_rated_sales": float(inv.zero_rated_sales),
        "sspt_sales": float(inv.sspt_sales),
        "percentage_tax_amount": float(inv.percentage_tax_amount),
        "subtotal": float(inv.subtotal),
        "total_amount": float(inv.total_amount),
        "discount_type": inv.discount_type,
        "discount_id_no": inv.discount_id_no,
        "discount_beneficiary_name": inv.discount_beneficiary_name,
        "discount_beneficiary_tin": inv.discount_beneficiary_tin,
        "discount_amount": float(inv.discount_amount),
        "status": inv.status,
        "voided_at": to_business_iso(inv.voided_at),
        "voided_reason": inv.voided_reason,
        "voided_by": inv.voided_by,
        "reprint_count": inv.reprint_count or 0,
        "last_reprinted_at": to_business_iso(inv.last_reprinted_at),
        "items": [
            {
                "product_id": item.product_id,
                "description": item.description,
                "quantity": float(item.quantity),
                "unit_cost": float(item.unit_cost),
                "line_total": float(item.line_total),
                "tax_line_classification": item.tax_line_classification,
            }
            for item in inv.items
        ],
    }


@invoices_bp.route("/", methods=["POST"])
@jwt_required()
def create_invoice():
    user_id = int(get_jwt_identity())
    user = User.query.filter_by(id=user_id).with_for_update().first()
    data = request.get_json() or {}

    cart = data.get("items", [])
    if not cart:
        return jsonify({"error": "Cart is empty"}), 400
    
    today = business_today(user.business_day_cutoff)
    existing_z = ZReading.query.filter_by(user_id=user_id, business_date=today).first()
    if existing_z:
        return jsonify({
            "error": "Z-Reading has already been generated for today. No further transactions can be recorded for this business date.",
            "z_reading_id": existing_z.id,
            "generated_at": to_business_iso(existing_z.generated_at)
        }), 409

    # --- Step 1: validate stock + resolve line items ---
    quantities_by_product = {}
    for entry in cart:
        try:
            product_id = int(entry["product_id"])
            quantity = Decimal(str(entry["quantity"]))
        except (KeyError, TypeError, ValueError):
            return jsonify({"error": "Invalid cart item"}), 400
        quantities_by_product[product_id] = quantities_by_product.get(product_id, Decimal("0")) + quantity

    line_items = []
    products_to_update = []

    for product_id, qty in quantities_by_product.items():
        product = Product.query.filter_by(
            id=product_id, user_id=user_id, is_archived=False
        ).with_for_update().first()
        if not product:
            return jsonify({"error": f"Product {product_id} not found"}), 404

        if qty <= 0:
            return jsonify({"error": "Quantity must be greater than zero"}), 400

        if product.stock_quantity < qty:
            return jsonify({
                "error": f"Insufficient stock for {product.name}. Available: {product.stock_quantity}{product.unit}"
            }), 400

        unit_cost = Decimal(str(product.price))
        line_total = (unit_cost * qty).quantize(Decimal("0.01"))
        tax_cls = resolve_tax_classification(product.tax_classification, user.vat_status)

        line_items.append({
            "product_id": product.id,
            "description": product.name + (f" ({product.cut_type})" if product.cut_type else ""),
            "quantity": qty,
            "unit_cost": unit_cost,
            "line_total": line_total,
            "tax_line_classification": tax_cls,
        })
        products_to_update.append((product, qty))

    # --- Step 2: compute totals ---
    totals = compute_invoice_totals(
        [{"line_total": li["line_total"], "tax_line_classification": li["tax_line_classification"]} for li in line_items],
        user.vat_status,
    )

    try:
        discount_amount = Decimal(str(data.get("discount_amount", 0)))
    except Exception:
        return jsonify({"error": "Invalid discount amount"}), 400
    if discount_amount < 0:
        return jsonify({"error": "Discount amount cannot be negative"}), 400
    discount_type = data.get("discount_type") or None
    allowed_discount_types = {"senior_citizen", "pwd", "naac", "solo_parent"}
    if discount_amount > 0:
        if discount_type not in allowed_discount_types:
            return jsonify({"error": "Select a valid statutory discount type"}), 400
        if not (data.get("discount_id_no") or "").strip():
            return jsonify({"error": "Discount beneficiary ID number is required"}), 400
        if not (data.get("discount_beneficiary_name") or "").strip():
            return jsonify({"error": "Discount beneficiary name is required"}), 400
    total_amount = totals["total_amount"] - discount_amount
    if total_amount < 0:
        return jsonify({"error": "Discount cannot exceed the invoice total"}), 400

    payment_mode = data.get("payment_mode", "cash")
    if payment_mode != "cash":
        return jsonify({"error": "Only cash payment is currently supported"}), 400
    try:
        cash_tendered = Decimal(str(data.get("cash_tendered", total_amount)))
    except Exception:
        return jsonify({"error": "Invalid cash tendered amount"}), 400
    if cash_tendered < total_amount:
        return jsonify({"error": "Cash tendered is less than the amount due"}), 400
    change_amount = cash_tendered - total_amount

    # --- Step 3: generate sequential invoice number ---
    invoice_number = user.next_invoice_number()

    # --- Step 4: create invoice record ---
    invoice = Invoice(
        user_id=user_id,
        invoice_number=invoice_number,
        invoice_type=data.get("invoice_type", "cash"),
        date_time=utc_now_naive(),
        cashier_name=user.email,
        payment_mode="cash",
        cash_tendered=cash_tendered,
        change_amount=change_amount,
        buyer_name=(data.get("buyer_name") or "").strip() or None,
        buyer_address=(data.get("buyer_address") or "").strip() or None,
        buyer_tin=(data.get("buyer_tin") or "").strip() or None,
        buyer_business_style=(data.get("buyer_business_style") or "").strip() or None,
        seller_snapshot_json=canonical_json(seller_snapshot(user)),
        vatable_sales=totals["vatable_sales"],
        vat_amount=totals["vat_amount"],
        vat_exempt_sales=totals["vat_exempt_sales"],
        zero_rated_sales=totals["zero_rated_sales"],
        sspt_sales=totals["sspt_sales"],
        percentage_tax_amount=totals["percentage_tax_amount"],
        subtotal=totals["subtotal"],
        total_amount=total_amount,
        discount_type=discount_type,
        discount_id_no=(data.get("discount_id_no") or "").strip() or None,
        discount_beneficiary_name=(data.get("discount_beneficiary_name") or "").strip() or None,
        discount_beneficiary_tin=(data.get("discount_beneficiary_tin") or "").strip() or None,
        discount_amount=discount_amount,
        status="active",
    )
    db.session.add(invoice)
    db.session.flush()  # get invoice.id

    # --- Step 5: create invoice items + deduct stock ---
    for li in line_items:
        item = InvoiceItem(
            invoice_id=invoice.id,
            product_id=li["product_id"],
            description=li["description"],
            quantity=li["quantity"],
            unit_cost=li["unit_cost"],
            line_total=li["line_total"],
            tax_line_classification=li["tax_line_classification"],
        )
        db.session.add(item)

    for product, qty in products_to_update:
        before_stock = float(product.stock_quantity)
        product.stock_quantity -= qty

        add_audit_event(
            user_id=user_id,
            entity_type="product",
            entity_id=product.id,
            action="sale_deduction",
            before={"stock_quantity": before_stock},
            after={"stock_quantity": float(product.stock_quantity)},
        )

    db.session.flush()

    # --- Step 6: e-journal entry ---
    db.session.add(EJournalEntry(
        user_id=user_id,
        entry_date=business_today(user.business_day_cutoff),
        entry_type="invoice",
        reference_id=invoice.id,
        snapshot_json=json.dumps(_invoice_to_dict(invoice)),
    ))
    add_audit_event(
        user_id=user_id,
        entity_type="invoice",
        entity_id=invoice.id,
        action="create",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        after=_invoice_to_dict(invoice),
    )

    db.session.commit()

    return jsonify(_invoice_to_dict(invoice)), 201


@invoices_bp.route("/", methods=["GET"])
@jwt_required()
def list_invoices():
    user_id = int(get_jwt_identity())
    query = Invoice.query.filter_by(user_id=user_id)
    date_value = request.args.get("date")
    start_value = request.args.get("start")
    end_value = request.args.get("end")
    status = request.args.get("status")

    try:
        if date_value:
            selected_date = datetime.strptime(date_value, "%Y-%m-%d").date()
            user = User.query.get(user_id)
            start_dt, end_dt = business_day_utc_bounds(selected_date, user.business_day_cutoff)
            query = query.filter(Invoice.date_time >= start_dt, Invoice.date_time <= end_dt)
        else:
            if start_value:
                query = query.filter(Invoice.date_time >= datetime.fromisoformat(start_value))
            if end_value:
                query = query.filter(Invoice.date_time < datetime.fromisoformat(end_value))
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    if status in ("active", "voided"):
        query = query.filter_by(status=status)

    invoices = query.order_by(Invoice.date_time.desc()).all()
    return jsonify([_invoice_to_dict(inv) for inv in invoices]), 200


@invoices_bp.route("/export", methods=["GET"])
@jwt_required()
def export_invoices():
    user_id = int(get_jwt_identity())
    date_value = request.args.get("date")
    if not date_value:
        return jsonify({"error": "date is required"}), 400
    try:
        selected_date = datetime.strptime(date_value, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "date must use YYYY-MM-DD format"}), 400

    user = User.query.get(user_id)
    start_dt, end_dt = business_day_utc_bounds(selected_date, user.business_day_cutoff)
    invoices = Invoice.query.filter(
        Invoice.user_id == user_id,
        Invoice.date_time >= start_dt,
        Invoice.date_time <= end_dt,
    ).order_by(Invoice.date_time).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices and Voids"
    header_fill = PatternFill("solid", fgColor="7F1D2D")
    total_fill = PatternFill("solid", fgColor="F5E8EA")
    white_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency = '"PHP" #,##0.00'

    ws.merge_cells("A1:I1")
    ws["A1"] = user.business_name
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Invoice and Void Register - {selected_date.isoformat()}"
    ws["A2"].font = bold
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Invoice No.", "Time", "Status", "Type", "Items", "Subtotal", "Discount", "Total", "Void Reason"]
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(4, column, value)
        cell.fill = header_fill
        cell.font = white_bold
        cell.border = border

    for row, invoice in enumerate(invoices, start=5):
        values = [
            invoice.invoice_number,
            invoice.date_time.strftime("%H:%M:%S"),
            invoice.status.title(),
            invoice.invoice_type.title(),
            len(invoice.items),
            float(invoice.subtotal),
            float(invoice.discount_amount),
            float(invoice.total_amount),
            invoice.voided_reason or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row, column, value)
            cell.border = border
            if column in (6, 7, 8):
                cell.number_format = currency

    total_row = 5 + len(invoices)
    ws.cell(total_row, 1, "ACTIVE SALES TOTAL").font = bold
    active_total = sum((invoice.total_amount for invoice in invoices if invoice.status == "active"), Decimal("0"))
    ws.cell(total_row, 8, float(active_total)).number_format = currency
    ws.cell(total_row + 1, 1, "VOID COUNT").font = bold
    ws.cell(total_row + 1, 8, sum(1 for invoice in invoices if invoice.status == "voided"))
    for row in (total_row, total_row + 1):
        for column in range(1, 10):
            ws.cell(row, column).border = border
            ws.cell(row, column).fill = total_fill

    widths = (16, 12, 12, 12, 10, 16, 16, 16, 32)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    items_ws = wb.create_sheet("Invoice Items")
    item_headers = ["Invoice No.", "Status", "Time", "Product", "Quantity", "Unit Price", "Line Total", "Tax Classification"]
    for column, value in enumerate(item_headers, start=1):
        cell = items_ws.cell(1, column, value)
        cell.fill = header_fill
        cell.font = white_bold
        cell.border = border
    item_row = 2
    for invoice in invoices:
        for item in invoice.items:
            values = [
                invoice.invoice_number,
                invoice.status.title(),
                invoice.date_time.strftime("%H:%M:%S"),
                item.description,
                float(item.quantity),
                float(item.unit_cost),
                float(item.line_total),
                item.tax_line_classification.replace("_", " ").title(),
            ]
            for column, value in enumerate(values, start=1):
                cell = items_ws.cell(item_row, column, value)
                cell.border = border
                if column in (6, 7):
                    cell.number_format = currency
            item_row += 1
    item_widths = (16, 12, 12, 32, 12, 16, 16, 22)
    for column, width in enumerate(item_widths, start=1):
        items_ws.column_dimensions[get_column_letter(column)].width = width
    items_ws.freeze_panes = "A2"
    items_ws.sheet_view.showGridLines = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    add_audit_event(
        user_id=user_id,
        entity_type="invoice_register",
        entity_id=int(selected_date.strftime("%Y%m%d")),
        action="export",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        after={"format": "xlsx", "business_date": selected_date},
    )
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"invoices-{selected_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@invoices_bp.route("/<int:invoice_id>", methods=["GET"])
@jwt_required()
def get_invoice(invoice_id):
    user_id = int(get_jwt_identity())
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=user_id).first()
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404
    return jsonify(_invoice_to_dict(invoice)), 200


@invoices_bp.route("/<int:invoice_id>/void", methods=["POST"])
@jwt_required()
def void_invoice(invoice_id):
    user_id = int(get_jwt_identity())
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=user_id).first()
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404

    if invoice.status == "voided":
        return jsonify({"error": "Invoice already voided"}), 400
    invoice_business_date = to_business_iso(invoice.date_time)[:10]
    existing_z = ZReading.query.filter_by(
        user_id=user_id,
        business_date=datetime.strptime(invoice_business_date, "%Y-%m-%d").date(),
    ).first()
    if existing_z:
        return jsonify({
            "error": "This business date is closed by a Z-Reading. Voids and other adjustments are no longer allowed."
        }), 409

    data = request.get_json() or {}
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"error": "A void reason is required"}), 400
    user = User.query.get(user_id)
    before = _invoice_to_dict(invoice)

    # Restore stock for each item
    for item in invoice.items:
        if item.product_id:
            product = Product.query.get(item.product_id)
            if product:
                before_stock = float(product.stock_quantity)
                product.stock_quantity += item.quantity
                add_audit_event(
                    user_id=user_id,
                    entity_type="product",
                    entity_id=product.id,
                    action="void_restock",
                    before={"stock_quantity": before_stock},
                    after={"stock_quantity": float(product.stock_quantity)},
                )

    invoice.status = "voided"
    invoice.voided_at = utc_now_naive()
    invoice.voided_reason = reason
    invoice.voided_by = user.email

    db.session.add(EJournalEntry(
        user_id=user_id,
        entry_date=business_today(user.business_day_cutoff),
        entry_type="void",
        reference_id=invoice.id,
        snapshot_json=canonical_json(_invoice_to_dict(invoice)),
    ))
    add_audit_event(
        user_id=user_id,
        entity_type="invoice",
        entity_id=invoice.id,
        action="void",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        before=before,
        after=_invoice_to_dict(invoice),
    )

    db.session.commit()
    return jsonify({"message": "Invoice voided", "invoice": _invoice_to_dict(invoice)}), 200


def _invoice_html(invoice: Invoice, reprint: bool) -> str:
    data = _invoice_to_dict(invoice)
    seller = data["seller"]
    vat_label = "VAT REG TIN" if seller.get("vat_status") == "vat" else "NON-VAT REG TIN"
    branch_code = seller.get("branch_code") or "00000"
    item_rows = "".join(
        "<tr>"
        f"<td>{escape(item['description'])}</td>"
        f"<td>{item['quantity']:.3f}</td>"
        f"<td>{item['unit_cost']:,.2f}</td>"
        f"<td>{item['line_total']:,.2f}</td>"
        "</tr>"
        for item in data["items"]
    )
    reprint_line = (
        f"<div class='reprint'>REPRINT - {escape(to_business_iso(invoice.last_reprinted_at) or '')}</div>"
        if reprint
        else ""
    )
    buyer = ""
    if invoice.buyer_name or invoice.buyer_tin:
        buyer = (
            "<section><strong>BUYER</strong><br>"
            f"{escape(invoice.buyer_name or '')}<br>"
            f"TIN: {escape(invoice.buyer_tin or '')}<br>"
            f"{escape(invoice.buyer_address or '')}</section>"
        )
    discount_section = ""
    if invoice.discount_amount and invoice.discount_amount > 0:
        discount_section = (
            "<section><strong>STATUTORY DISCOUNT</strong><br>"
            f"Type: {escape((invoice.discount_type or '').replace('_', ' ').upper())}<br>"
            f"Beneficiary: {escape(invoice.discount_beneficiary_name or '')}<br>"
            f"TIN: {escape(invoice.discount_beneficiary_tin or '')}<br>"
            f"ID No.: {escape(invoice.discount_id_no or '')}<br><br>"
            "Signature: ______________________________</section>"
        )
    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Invoice {escape(data['display_invoice_number'])}</title>
<style>
body{{font:12px Arial,sans-serif;width:76mm;margin:0 auto;color:#111}} h1,p{{margin:2px 0}}
header,footer,.center{{text-align:center}} table{{width:100%;border-collapse:collapse;margin:8px 0}}
th,td{{padding:3px 1px;border-bottom:1px dashed #777;text-align:right}} th:first-child,td:first-child{{text-align:left}}
.totals div{{display:flex;justify-content:space-between}} .grand{{font-size:16px;font-weight:700}}
.reprint{{font-size:18px;font-weight:700;text-align:center;border:2px solid #111;margin:6px 0}}
section{{margin:8px 0}} @media print{{button{{display:none}}}}
</style></head><body>
<header><h1>{escape(seller.get('registered_name') or '')}</h1>
<p>{escape(seller.get('business_name') or '')}</p>
<p>{escape(seller.get('business_address') or '')}</p>
<p>{vat_label}: {escape(seller.get('tin') or '')}-{escape(branch_code)}</p>
<p>MIN: {escape(seller.get('machine_identification_number') or '')}</p>
<p>Serial/License: {escape(seller.get('machine_serial_number') or seller.get('software_license_number') or '')}</p>
<h1>INVOICE</h1><p>No. {escape(data['display_invoice_number'])}</p>
<p>{escape(to_business_iso(invoice.date_time) or '')}</p></header>
{reprint_line}{buyer}{discount_section}
<table><thead><tr><th>Description</th><th>Qty</th><th>Unit</th><th>Amount</th></tr></thead>
<tbody>{item_rows}</tbody></table>
<div class="totals">
<div><span>VATable Sales</span><span>PHP {data['vatable_sales']:,.2f}</span></div>
<div><span>VAT-Exempt Sales</span><span>PHP {data['vat_exempt_sales']:,.2f}</span></div>
<div><span>Zero-Rated Sales</span><span>PHP {data['zero_rated_sales']:,.2f}</span></div>
<div><span>VAT Amount</span><span>PHP {data['vat_amount']:,.2f}</span></div>
<div><span>SSPT Sales</span><span>PHP {data['sspt_sales']:,.2f}</span></div>
<div><span>Discount</span><span>PHP {data['discount_amount']:,.2f}</span></div>
<div class="grand"><span>TOTAL</span><span>PHP {data['total_amount']:,.2f}</span></div>
<div><span>Cash</span><span>PHP {data['cash_tendered']:,.2f}</span></div>
<div><span>Change</span><span>PHP {data['change_amount']:,.2f}</span></div>
</div>
<footer><p>Cashier: {escape(invoice.cashier_name or '')}</p>
<p>Accreditation: {escape(seller.get('accreditation_number') or '')}</p>
<p>Accreditation issued: {escape(str(seller.get('accreditation_date_issued') or ''))} | valid until: {escape(str(seller.get('accreditation_valid_until') or ''))}</p>
<p>PTU: {escape(seller.get('ptu_number') or '')}</p>
<p>PTU issued: {escape(str(seller.get('ptu_date_issued') or ''))}</p>
<p>Supplier: {escape(seller.get('accredited_supplier_name') or '')}</p>
<p>{escape(seller.get('accredited_supplier_address') or '')}</p>
<p>Supplier TIN: {escape(seller.get('accredited_supplier_tin') or '')}</p>
<p>Software version: {escape(seller.get('software_version') or '')}</p>
<button onclick="window.print()">Print</button></footer></body></html>"""


@invoices_bp.route("/<int:invoice_id>/print", methods=["GET"])
@jwt_required()
def print_invoice(invoice_id):
    user_id = int(get_jwt_identity())
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=user_id).first()
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404
    return Response(_invoice_html(invoice, False), mimetype="text/html")


@invoices_bp.route("/<int:invoice_id>/reprint", methods=["POST"])
@jwt_required()
def reprint_invoice(invoice_id):
    user_id = int(get_jwt_identity())
    invoice = Invoice.query.filter_by(id=invoice_id, user_id=user_id).with_for_update().first()
    if not invoice:
        return jsonify({"error": "Invoice not found"}), 404
    user = User.query.get(user_id)
    invoice.reprint_count = (invoice.reprint_count or 0) + 1
    invoice.last_reprinted_at = utc_now_naive()
    add_audit_event(
        user_id=user_id,
        entity_type="invoice",
        entity_id=invoice.id,
        action="reprint",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        after={
            "invoice_number": invoice.invoice_number,
            "reprint_count": invoice.reprint_count,
            "reprinted_at": invoice.last_reprinted_at,
        },
    )
    db.session.commit()
    return Response(_invoice_html(invoice, True), mimetype="text/html")
