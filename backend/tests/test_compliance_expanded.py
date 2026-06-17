from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app import create_app
from models.audit_log import AuditLog
from models.db import db
from models.invoice import Invoice
from models.product import Product
from models.user import User


def _app():
    app = create_app({
        "TESTING": True,
        "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
        "SECRET_KEY": "test-app-secret-that-is-at-least-32-bytes",
        "JWT_SECRET_KEY": "test-jwt-secret-that-is-at-least-32-bytes",
    })
    with app.app_context():
        db.drop_all()
        db.create_all()
    return app


def _json(response, status):
    assert response.status_code == status, response.get_json()
    return response.get_json()


def _register(client, email="owner@example.com", vat_status="non_vat"):
    registered = _json(client.post("/api/auth/register", json={
        "email": email,
        "password": "Strong1!",
        "business_name": "Compliance Store",
        "vat_status": vat_status,
    }), 201)
    return registered, {"Authorization": f"Bearer {registered['access_token']}"}


def _product(client, headers, name="Item", price="100.00", stock="10", tax_classification="standard"):
    return _json(client.post("/api/products/", headers=headers, json={
        "name": name,
        "category": "retail",
        "pricing_type": "fixed",
        "price": price,
        "stock_quantity": stock,
        "unit": "pcs",
        "tax_classification": tax_classification,
    }), 201)


def _invoice(client, headers, product, quantity=1, **extra):
    payload = {
        "items": [{"product_id": product["id"], "quantity": quantity}],
        "cash_tendered": extra.pop("cash_tendered", "1000.00"),
    }
    payload.update(extra)
    return _json(client.post("/api/invoices/", headers=headers, json=payload), 201)


def _set_invoice_time(app, invoice_id, utc_dt):
    with app.app_context():
        invoice = Invoice.query.get(invoice_id)
        invoice.date_time = utc_dt
        db.session.commit()


def _set_role(app, user_id, role):
    with app.app_context():
        user = User.query.get(user_id)
        user.role = role
        db.session.commit()


def test_z_reading_counters_accumulated_totals_and_business_date_boundaries():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="100.00", stock="10")

    first = _invoice(client, headers, product, quantity=1)
    _set_invoice_time(app, first["id"], datetime(2026, 6, 16, 2, 0, 0))
    z1 = _json(client.post("/api/readings/z", headers=headers, json={
        "password": "Strong1!",
        "date": "2026-06-16",
    }), 201)
    assert z1["z_counter"] == 1
    assert z1["business_date"] == "2026-06-16"
    assert z1["total_sales"] == 100.0
    assert z1["accumulated_grand_total"] == 100.0

    second = _invoice(client, headers, product, quantity=2)
    _set_invoice_time(app, second["id"], datetime(2026, 6, 17, 2, 0, 0))
    z2 = _json(client.post("/api/readings/z", headers=headers, json={
        "password": "Strong1!",
        "date": "2026-06-17",
    }), 201)
    assert z2["z_counter"] == 2
    assert z2["business_date"] == "2026-06-17"
    assert z2["total_sales"] == 200.0
    assert z2["accumulated_grand_total"] == 300.0

    duplicate = client.post("/api/readings/z", headers=headers, json={
        "password": "Strong1!",
        "date": "2026-06-17",
    })
    assert duplicate.status_code == 409


def test_z_reading_blocks_new_invoices_for_closed_current_business_date():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="50.00", stock="5")
    _invoice(client, headers, product, quantity=1)

    _json(client.post("/api/readings/z", headers=headers, json={"password": "Strong1!"}), 201)
    blocked = client.post("/api/invoices/", headers=headers, json={
        "items": [{"product_id": product["id"], "quantity": 1}],
        "cash_tendered": "100.00",
    })
    assert blocked.status_code == 409


def test_z_reading_uses_manila_business_date_for_utc_boundary_transaction():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="75.00", stock="5")
    invoice = _invoice(client, headers, product, quantity=1)
    _set_invoice_time(app, invoice["id"], datetime(2026, 6, 16, 23, 30, 0))

    z_reading = _json(client.post("/api/readings/z", headers=headers, json={
        "password": "Strong1!",
        "date": "2026-06-17",
    }), 201)
    assert z_reading["business_date"] == "2026-06-17"
    assert z_reading["total_sales"] == 75.0


def test_void_restores_stock_requires_reason_blocks_after_z_and_counts_in_z():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="20.00", stock="5")
    invoice = _invoice(client, headers, product, quantity=2)

    missing_reason = client.post(f"/api/invoices/{invoice['id']}/void", headers=headers, json={})
    assert missing_reason.status_code == 400

    voided = _json(client.post(f"/api/invoices/{invoice['id']}/void", headers=headers, json={
        "reason": "Wrong quantity",
    }), 200)
    assert voided["invoice"]["status"] == "voided"

    products = _json(client.get("/api/products/", headers=headers), 200)
    restored = next(item for item in products if item["id"] == product["id"])
    assert restored["stock_quantity"] == 5.0

    already_voided = client.post(f"/api/invoices/{invoice['id']}/void", headers=headers, json={
        "reason": "Duplicate void",
    })
    assert already_voided.status_code == 409

    product_after_z = _product(client, headers, name="Post Z Item", price="30.00", stock="5")
    old_invoice = _invoice(client, headers, product_after_z, quantity=1)
    _set_invoice_time(app, old_invoice["id"], datetime(2026, 6, 16, 23, 30, 0))

    z_reading = _json(client.post("/api/readings/z", headers=headers, json={"password": "Strong1!"}), 201)
    assert z_reading["void_count"] == 1

    blocked_void = client.post(f"/api/invoices/{old_invoice['id']}/void", headers=headers, json={
        "reason": "Closed day",
    })
    assert blocked_void.status_code == 409


def test_exports_use_manila_datetime_and_ejournal_contains_required_entry_types():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="80.00", stock="5")
    invoice = _invoice(client, headers, product, quantity=1)
    _set_invoice_time(app, invoice["id"], datetime(2026, 6, 16, 23, 30, 0))

    invoice_export = client.get("/api/invoices/export?date=2026-06-17", headers=headers)
    assert invoice_export.status_code == 200
    wb = load_workbook(BytesIO(invoice_export.data))
    assert wb["Invoices and Voids"]["B5"].value == "07:30:00"

    _json(client.post(f"/api/invoices/{invoice['id']}/void", headers=headers, json={
        "reason": "Customer cancelled",
    }), 200)
    _json(client.post("/api/readings/x", headers=headers, json={"date": "2026-06-17"}), 201)
    _json(client.post("/api/readings/z", headers=headers, json={
        "password": "Strong1!",
        "date": "2026-06-17",
    }), 201)

    ejournal = client.get("/api/readings/ejournal/export?date=2026-06-17", headers=headers)
    assert ejournal.status_code == 200
    body = ejournal.data.decode("utf-8")
    assert "[INVOICE]" in body
    assert "[VOID]" in body
    assert "[X-READING]" in body
    assert "[Z-READING]" in body

    with app.app_context():
        entry = AuditLog.query.first()
        entry.created_at = datetime(2026, 6, 16, 23, 30, 0)
        db.session.commit()

    audit_export = client.get("/api/compliance/audit-log/export?date=2026-06-17", headers=headers)
    assert audit_export.status_code == 200
    audit_wb = load_workbook(BytesIO(audit_export.data))
    assert audit_wb["Activity Log"]["B2"].value == "2026-06-17 07:30:00"


def test_naac_and_solo_parent_discounts_compute_without_vat_deduction():
    app = _app()
    client = app.test_client()
    _, headers = _register(client, email="discount-expanded@example.com", vat_status="vat")
    product = _product(client, headers, price="112.00", stock="10")

    naac = _invoice(client, headers, product, quantity=1, discount_type="naac", discount_id_no="NAAC-1")
    assert naac["discount_amount"] == 22.4
    assert naac["vat_deduction"] == 0.0
    assert naac["total_amount"] == 89.6

    solo_parent = _invoice(client, headers, product, quantity=1, discount_type="solo_parent", discount_id_no="SP-1")
    assert solo_parent["discount_amount"] == 11.2
    assert solo_parent["vat_deduction"] == 0.0
    assert solo_parent["total_amount"] == 100.8


def test_x_reading_uses_manila_business_date_for_utc_boundary_transaction():
    app = _app()
    client = app.test_client()
    _, headers = _register(client)
    product = _product(client, headers, price="60.00", stock="5")
    invoice = _invoice(client, headers, product, quantity=1)
    _set_invoice_time(app, invoice["id"], datetime(2026, 6, 16, 23, 30, 0))

    x_reading = _json(client.post("/api/readings/x", headers=headers, json={
        "date": "2026-06-17",
    }), 201)
    assert x_reading["shift_date"] == "2026-06-17"
    assert x_reading["total_sales"] == 60.0


def test_rbac_matrix_for_cashier_auditor_and_owner():
    app = _app()
    client = app.test_client()
    registered, headers = _register(client)
    product = _product(client, headers, price="40.00", stock="5")
    _invoice(client, headers, product, quantity=1)

    owner_z = client.post("/api/readings/z", headers=headers, json={"password": "Strong1!"})
    assert owner_z.status_code == 201

    cashier_registered, cashier_headers = _register(client, email="cashier-rbac@example.com")
    _set_role(app, cashier_registered["user"]["id"], "cashier")
    cashier_z = client.post("/api/readings/z", headers=cashier_headers, json={"password": "Strong1!"})
    assert cashier_z.status_code == 403
    cashier_edit = client.put(f"/api/products/{product['id']}", headers=cashier_headers, json={"name": "Nope"})
    assert cashier_edit.status_code == 403

    auditor_registered, auditor_headers = _register(client, email="auditor-rbac@example.com")
    _set_role(app, auditor_registered["user"]["id"], "auditor")
    auditor_create_invoice = client.post("/api/invoices/", headers=auditor_headers, json={
        "items": [{"product_id": product["id"], "quantity": 1}],
        "cash_tendered": "100.00",
    })
    assert auditor_create_invoice.status_code == 403
    auditor_list = client.get("/api/invoices/", headers=auditor_headers)
    assert auditor_list.status_code == 200
