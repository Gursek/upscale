import hashlib
from datetime import datetime
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import get_jwt_identity, jwt_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models.audit_log import AuditLog
from models.user import User
from services.compliance import add_audit_event, canonical_json, compliance_readiness
from models.db import db
from services.business_time import business_day_utc_bounds, to_business_datetime
from services.rbac import roles_required


compliance_bp = Blueprint("compliance", __name__)


@compliance_bp.route("/readiness", methods=["GET"])
@jwt_required()
@roles_required("owner", "auditor")
def readiness():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    result = compliance_readiness(user)
    result["software_version"] = user.software_version
    result["reset_counter"] = user.reset_counter or 0
    return jsonify(result), 200


@compliance_bp.route("/audit-integrity", methods=["GET"])
@jwt_required()
@roles_required("owner", "auditor")
def audit_integrity():
    user_id = int(get_jwt_identity())
    entries = AuditLog.query.filter_by(user_id=user_id).order_by(AuditLog.id).all()
    previous_hash = None
    checked = 0
    for entry in entries:
        if not entry.event_hash:
            continue
        payload = {
            "user_id": entry.user_id,
            "entity_type": entry.entity_type,
            "entity_id": entry.entity_id,
            "action": entry.action,
            "actor": entry.actor,
            "terminal_id": entry.terminal_id,
            "before": (
                __import__("json").loads(entry.before_state)
                if entry.before_state
                else None
            ),
            "after": (
                __import__("json").loads(entry.after_state)
                if entry.after_state
                else None
            ),
            "created_at": entry.created_at,
            "previous_hash": entry.previous_hash,
        }
        expected = hashlib.sha256(canonical_json(payload).encode("utf-8")).hexdigest()
        if entry.previous_hash != previous_hash or entry.event_hash != expected:
            return jsonify({
                "valid": False,
                "failed_event_id": entry.id,
                "checked_events": checked,
            }), 409
        previous_hash = entry.event_hash
        checked += 1
    return jsonify({"valid": True, "checked_events": checked}), 200


@compliance_bp.route("/audit-log/export", methods=["GET"])
@jwt_required()
@roles_required("owner", "auditor")
def export_audit_log():
    user_id = int(get_jwt_identity())
    date_value = request.args.get("date")
    query = AuditLog.query.filter_by(user_id=user_id)
    if date_value:
        try:
            selected_date = datetime.strptime(date_value, "%Y-%m-%d").date()
        except ValueError:
            return jsonify({"error": "date must use YYYY-MM-DD format"}), 400
        user = User.query.get(user_id)
        start_dt, end_dt = business_day_utc_bounds(selected_date, user.business_day_cutoff)
        query = query.filter(
            AuditLog.created_at >= start_dt,
            AuditLog.created_at <= end_dt,
        )
    entries = query.order_by(AuditLog.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Log"
    headers = [
        "Event ID", "Date/Time Manila", "Actor", "Terminal", "IP Address",
        "Entity", "Entity ID", "Action", "Before", "After",
        "Previous Hash", "Event Hash",
    ]
    fill = PatternFill("solid", fgColor="7F1D2D")
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(1, column, value)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row, entry in enumerate(entries, start=2):
        values = [
            entry.id,
            to_business_datetime(entry.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            entry.actor or "",
            entry.terminal_id or "",
            entry.ip_address or "",
            entry.entity_type,
            entry.entity_id,
            entry.action,
            entry.before_state or "",
            entry.after_state or "",
            entry.previous_hash or "",
            entry.event_hash or "",
        ]
        for column, value in enumerate(values, start=1):
            ws.cell(row, column, value)
    widths = (10, 22, 28, 18, 16, 18, 12, 18, 45, 45, 68, 68)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = date_value or "all"
    add_audit_event(
        user_id=user_id,
        entity_type="activity_log",
        entity_id=int(date_value.replace("-", "")) if date_value else 0,
        action="export",
        after={"format": "xlsx", "business_date": date_value or "all"},
    )
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"activity-log-{suffix}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
