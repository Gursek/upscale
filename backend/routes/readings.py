from flask import Blueprint, request, jsonify, send_file, Response
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash

from models.db import db
from models.ejournal import EJournalEntry, XReading, ZReading
from models.invoice import Invoice, InvoiceItem
from models.product import Product
from models.user import User
from services.readings import generate_x_reading, generate_z_reading
from services.business_time import (
    business_day_utc_bounds,
    business_today,
    to_business_datetime,
    to_business_iso,
)
from services.compliance import add_audit_event

readings_bp = Blueprint("readings", __name__)


def _report_breakdown(user_id, report_date):
    user = User.query.get(user_id)
    start_dt, end_dt = business_day_utc_bounds(report_date, user.business_day_cutoff)
    invoices = Invoice.query.filter(
        Invoice.user_id == user_id,
        Invoice.date_time >= start_dt,
        Invoice.date_time <= end_dt,
    ).all()
    active = [invoice for invoice in invoices if invoice.status == "active"]
    return {
        "sspt_sales": float(sum((invoice.sspt_sales for invoice in active), Decimal("0"))),
        "percentage_tax_amount": float(sum((invoice.percentage_tax_amount for invoice in active), Decimal("0"))),
        "discount_amount": float(sum((invoice.discount_amount for invoice in active), Decimal("0"))),
        "net_total": float(sum((invoice.total_amount for invoice in active), Decimal("0"))),
    }


def _x_to_dict(r: XReading) -> dict:
    result = {
        "id": r.id,
        "shift_date": r.shift_date.isoformat(),
        "generated_at": to_business_iso(r.generated_at),
        "total_sales": float(r.total_sales),
        "vatable_sales": float(r.vatable_sales),
        "vat_exempt_sales": float(r.vat_exempt_sales),
        "zero_rated_sales": float(r.zero_rated_sales),
        "vat_amount": float(r.vat_amount),
        "transaction_count": r.transaction_count,
        "void_count": r.void_count,
        "starting_invoice_no": r.starting_invoice_no,
        "ending_invoice_no": r.ending_invoice_no,
        "cash_sales": float(r.cash_sales or 0),
        "reset_counter": r.reset_counter or 0,
    }
    result.update(_report_breakdown(r.user_id, r.shift_date))
    return result


def _z_to_dict(r: ZReading) -> dict:
    result = {
        "id": r.id,
        "business_date": r.business_date.isoformat(),
        "generated_at": to_business_iso(r.generated_at),
        "total_sales": float(r.total_sales),
        "vatable_sales": float(r.vatable_sales),
        "vat_exempt_sales": float(r.vat_exempt_sales),
        "zero_rated_sales": float(r.zero_rated_sales),
        "vat_amount": float(r.vat_amount),
        "transaction_count": r.transaction_count,
        "void_count": r.void_count,
        "accumulated_grand_total": float(r.accumulated_grand_total),
        "z_counter": r.z_counter,
        "starting_invoice_no": r.starting_invoice_no,
        "ending_invoice_no": r.ending_invoice_no,
        "cash_sales": float(r.cash_sales or 0),
        "reset_counter": r.reset_counter or 0,
    }
    result.update(_report_breakdown(r.user_id, r.business_date))
    return result


@readings_bp.route("/x", methods=["POST"])
@jwt_required()
def create_x_reading():
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    user = User.query.get(user_id)

    shift_date = business_today(user.business_day_cutoff)
    if data.get("date"):
        shift_date = datetime.strptime(data["date"], "%Y-%m-%d").date()

    reading, _ = generate_x_reading(user_id, shift_date)

    db.session.add(EJournalEntry(
        user_id=user_id,
        entry_date=business_today(user.business_day_cutoff),
        entry_type="x_reading",
        reference_id=reading.id,
        snapshot_json=json.dumps(_x_to_dict(reading)),
    ))
    add_audit_event(
        user_id=user_id,
        entity_type="x_reading",
        entity_id=reading.id,
        action="generate",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        after=_x_to_dict(reading),
    )
    db.session.commit()

    return jsonify(_x_to_dict(reading)), 201


@readings_bp.route("/z", methods=["POST"])
@jwt_required()
def create_z_reading():
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    user = User.query.get(user_id)

    if not user or not check_password_hash(user.password_hash, data.get("password", "")):
        return jsonify({"error": "Password confirmation failed"}), 401

    business_date = business_today(user.business_day_cutoff)
    if data.get("date"):
        business_date = datetime.strptime(data["date"], "%Y-%m-%d").date()

    # Prevent duplicate Z-reading for the same business date
    existing = ZReading.query.filter_by(user_id=user_id, business_date=business_date).first()
    if existing:
        return jsonify({"error": f"Z-Reading already generated for {business_date}"}), 409

    reading, _ = generate_z_reading(user_id, business_date)

    db.session.add(EJournalEntry(
        user_id=user_id,
        entry_date=business_today(user.business_day_cutoff),
        entry_type="z_reading",
        reference_id=reading.id,
        snapshot_json=json.dumps(_z_to_dict(reading)),
    ))
    add_audit_event(
        user_id=user_id,
        entity_type="z_reading",
        entity_id=reading.id,
        action="generate",
        actor=user.email,
        terminal_id=user.machine_identification_number,
        after=_z_to_dict(reading),
    )
    db.session.commit()

    return jsonify(_z_to_dict(reading)), 201


@readings_bp.route("/x", methods=["GET"])
@jwt_required()
def list_x_readings():
    user_id = int(get_jwt_identity())
    readings = XReading.query.filter_by(user_id=user_id).order_by(XReading.generated_at.desc()).all()
    return jsonify([_x_to_dict(r) for r in readings]), 200


@readings_bp.route("/z", methods=["GET"])
@jwt_required()
def list_z_readings():
    user_id = int(get_jwt_identity())
    readings = ZReading.query.filter_by(user_id=user_id).order_by(ZReading.generated_at.desc()).all()
    return jsonify([_z_to_dict(r) for r in readings]), 200


def _category_breakdown(user_id, business_date):
    user = User.query.get(user_id)
    start_dt, end_dt = business_day_utc_bounds(business_date, user.business_day_cutoff)
    rows = db.session.query(
        Product.category,
        db.func.sum(InvoiceItem.line_total),
    ).join(
        Invoice, Invoice.id == InvoiceItem.invoice_id
    ).outerjoin(
        Product, Product.id == InvoiceItem.product_id
    ).filter(
        Invoice.user_id == user_id,
        Invoice.status == "active",
        Invoice.date_time >= start_dt,
        Invoice.date_time <= end_dt,
    ).group_by(Product.category).all()

    return [(category or "uncategorized", Decimal(str(total or 0))) for category, total in rows]


@readings_bp.route("/z/<int:reading_id>/export", methods=["GET"])
@jwt_required()
def export_z_reading(reading_id):
    user_id = int(get_jwt_identity())
    reading = ZReading.query.filter_by(id=reading_id, user_id=user_id).first()
    if not reading:
        return jsonify({"error": "Z-Reading not found"}), 404

    user = User.query.get(user_id)
    categories = _category_breakdown(user_id, reading.business_date)
    wb = Workbook()
    ws = wb.active
    ws.title = "Z-Reading"

    dark_fill = PatternFill("solid", fgColor="7F1D2D")
    light_fill = PatternFill("solid", fgColor="F5E8EA")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_format = "₱#,##0.00"

    ws.merge_cells("A1:D1")
    ws["A1"] = user.business_name
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = user.business_address or ""
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:D3")
    ws["A3"] = f"TIN: {user.tin or 'Not provided'} | {user.vat_status.replace('_', '-').upper()}"
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A5:D5")
    ws["A5"] = "Z-READING REPORT"
    ws["A5"].font = Font(size=13, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A6:D6")
    ws["A6"] = f"Date: {reading.business_date.isoformat()} | Z-Counter: {reading.z_counter}"
    ws["A6"].alignment = Alignment(horizontal="center")

    breakdown = _report_breakdown(user_id, reading.business_date)
    summary = [
        ("Generated At", to_business_datetime(reading.generated_at).strftime("%Y-%m-%d %H:%M:%S"), "", ""),
        ("Starting Invoice", reading.starting_invoice_no or "-", "Ending Invoice", reading.ending_invoice_no or "-"),
        ("Transactions", reading.transaction_count, "Voids", reading.void_count),
        ("Payment Mode", "Cash", "Reset Counter", reading.reset_counter or 0),
    ]
    for row_index, values in enumerate(summary, start=8):
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row_index, column_index, value)
            cell.border = border
            if column_index in (1, 3):
                cell.font = bold_font
                cell.fill = light_fill

    ws["A12"] = "Sales Breakdown"
    ws["A12"].font = white_font
    ws["A12"].fill = dark_fill
    ws.merge_cells("A12:D12")
    sales_rows = [
        ("Gross Sales", reading.total_sales),
        ("VATable Sales", reading.vatable_sales),
        ("VAT-Exempt Sales", reading.vat_exempt_sales),
        ("Zero-Rated Sales", reading.zero_rated_sales),
        ("SSPT Sales", breakdown["sspt_sales"]),
        ("VAT Amount", reading.vat_amount),
        ("Percentage Tax", breakdown["percentage_tax_amount"]),
        ("Discounts", breakdown["discount_amount"]),
        ("Net Total", breakdown["net_total"]),
        ("Cash Sales", reading.cash_sales or 0),
    ]
    for row_index, (label, amount) in enumerate(sales_rows, start=13):
        ws.cell(row_index, 1, label).font = bold_font
        ws.cell(row_index, 4, float(amount)).number_format = currency_format
        for column_index in range(1, 5):
            ws.cell(row_index, column_index).border = border

    category_start = 23
    ws.merge_cells(start_row=category_start, start_column=1, end_row=category_start, end_column=4)
    ws.cell(category_start, 1, "Sales by Category").font = white_font
    ws.cell(category_start, 1).fill = dark_fill
    for row_index, (category, total) in enumerate(categories, start=category_start + 1):
        ws.cell(row_index, 1, category.replace("_", " ").title())
        ws.cell(row_index, 4, float(total)).number_format = currency_format
        for column_index in range(1, 5):
            ws.cell(row_index, column_index).border = border

    total_row = category_start + 1 + len(categories)
    ws.cell(total_row, 1, "Category Total").font = bold_font
    ws.cell(total_row, 4, f"=SUM(D{category_start + 1}:D{total_row - 1})")
    ws.cell(total_row, 4).font = bold_font
    ws.cell(total_row, 4).number_format = currency_format
    for column_index in range(1, 5):
        ws.cell(total_row, column_index).border = border
        ws.cell(total_row, column_index).fill = light_fill

    for column_index, width in enumerate((24, 18, 24, 18), start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width
    footer_row = total_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=3)
    ws.cell(footer_row, 1, "ACCUMULATED GRAND TOTAL")
    ws.cell(footer_row, 1).font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(footer_row, 1).fill = dark_fill
    ws.cell(footer_row, 4, float(reading.accumulated_grand_total))
    ws.cell(footer_row, 4).number_format = "₱#,##0.00"
    ws.cell(footer_row, 4).font = Font(size=12, bold=True, color="FFFFFF")
    ws.cell(footer_row, 4).fill = dark_fill
    for column_index in range(1, 5):
        ws.cell(footer_row, column_index).border = border
    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:D{footer_row}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    add_audit_event(
        user_id=user_id,
        entity_type="z_reading",
        entity_id=reading.id,
        action="export",
        after={"format": "xlsx", "business_date": reading.business_date},
    )
    db.session.commit()
    filename = f"z-reading-{reading.business_date.isoformat()}-{reading.z_counter}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@readings_bp.route("/ejournal/export", methods=["GET"])
@jwt_required()
def export_ejournal():
    user_id = int(get_jwt_identity())
    requested_date = request.args.get("date")
    try:
        user = User.query.get(user_id)
        export_date = (
            datetime.strptime(requested_date, "%Y-%m-%d").date()
            if requested_date
            else business_today(user.business_day_cutoff)
        )
    except ValueError:
        return jsonify({"error": "date must use YYYY-MM-DD format"}), 400

    user = User.query.get(user_id)
    entries = EJournalEntry.query.filter_by(
        user_id=user_id, entry_date=export_date
    ).order_by(EJournalEntry.created_at, EJournalEntry.id).all()

    lines = [
        "=== E-JOURNAL ===",
        f"Business: {user.business_name}",
        f"TIN: {user.tin or ''}",
        f"Date: {export_date.isoformat()}",
        "=====================================",
    ]
    for entry in entries:
        snapshot = json.loads(entry.snapshot_json)
        entry_time = to_business_datetime(entry.created_at).strftime("%Y-%m-%d %H:%M:%S")
        if entry.entry_type == "invoice":
            lines.append(f"[INVOICE] #{snapshot.get('invoice_number', '')} {entry_time}")
            seller = snapshot.get("seller") or {}
            lines.append(
                f"  Seller: {seller.get('registered_name', '')} | "
                f"TIN: {seller.get('tin', '')}-{seller.get('branch_code', '')} | "
                f"MIN: {seller.get('machine_identification_number', '')}"
            )
            if snapshot.get("buyer_name") or snapshot.get("buyer_tin"):
                lines.append(
                    f"  Buyer: {snapshot.get('buyer_name') or ''} | "
                    f"TIN: {snapshot.get('buyer_tin') or ''} | "
                    f"Address: {snapshot.get('buyer_address') or ''}"
                )
            for item in snapshot.get("items", []):
                product = Product.query.get(item.get("product_id")) if item.get("product_id") else None
                unit = product.unit if product else "unit"
                lines.append(
                    f"  {item.get('description', '')} - {float(item.get('quantity', 0)):.3f}{unit} x "
                    f"₱{float(item.get('unit_cost', 0)):,.2f} = ₱{float(item.get('line_total', 0)):,.2f} "
                    f"[{str(item.get('tax_line_classification', '')).replace('_', '-').upper()}]"
                )
            lines.append(f"  TOTAL: ₱{float(snapshot.get('total_amount', 0)):,.2f}")
            lines.append(
                f"  VATable: ₱{float(snapshot.get('vatable_sales', 0)):,.2f} | "
                f"VAT: ₱{float(snapshot.get('vat_amount', 0)):,.2f} | "
                f"VAT-Exempt: ₱{float(snapshot.get('vat_exempt_sales', 0)):,.2f} | "
                f"Zero-Rated: ₱{float(snapshot.get('zero_rated_sales', 0)):,.2f} | "
                f"SSPT: ₱{float(snapshot.get('sspt_sales', 0)):,.2f}"
            )
            lines.append(
                f"  Payment: CASH ₱{float(snapshot.get('cash_tendered', 0)):,.2f} | "
                f"Change: ₱{float(snapshot.get('change_amount', 0)):,.2f} | "
                f"Cashier: {snapshot.get('cashier_name') or ''}"
            )
            if snapshot.get("discount_amount", 0):
                lines.append(
                    f"  Discount: {snapshot.get('discount_type') or ''} | "
                    f"Beneficiary: {snapshot.get('discount_beneficiary_name') or ''} | "
                    f"ID: {snapshot.get('discount_id_no') or ''} | "
                    f"Amount: ₱{float(snapshot.get('discount_amount', 0)):,.2f}"
                )
        elif entry.entry_type == "void":
            lines.append(f"[VOID] #{snapshot.get('invoice_number', '')} {entry_time}")
            lines.append(f"  Reason: {snapshot.get('voided_reason') or snapshot.get('reason') or 'Not specified'}")
            lines.append(f"  Voided By: {snapshot.get('voided_by') or ''}")
        elif entry.entry_type == "x_reading":
            lines.append(f"[X-READING] {entry_time}")
            lines.append(
                f"  Total Sales: ₱{float(snapshot.get('total_sales', 0)):,.2f} | "
                f"Cash: ₱{float(snapshot.get('cash_sales', 0)):,.2f} | "
                f"Transactions: {snapshot.get('transaction_count', 0)} | Voids: {snapshot.get('void_count', 0)}"
            )
        elif entry.entry_type == "z_reading":
            lines.append(f"[Z-READING] #{snapshot.get('z_counter', '')} {entry_time}")
            lines.append(
                f"  Total Sales: ₱{float(snapshot.get('total_sales', 0)):,.2f} | "
                f"Cash: ₱{float(snapshot.get('cash_sales', 0)):,.2f} | "
                f"Accumulated: ₱{float(snapshot.get('accumulated_grand_total', 0)):,.2f} | "
                f"Reset Counter: {snapshot.get('reset_counter', 0)}"
            )
        else:
            lines.append(f"[{entry.entry_type.upper()}] {entry_time}")
        lines.append("-------------------------------------")
    lines.append("=====================================")

    body = "\r\n".join(lines) + "\r\n"
    add_audit_event(
        user_id=user_id,
        entity_type="e_journal",
        entity_id=int(export_date.strftime("%Y%m%d")),
        action="export",
        after={"format": "txt", "business_date": export_date},
    )
    db.session.commit()
    return Response(
        body,
        mimetype="text/plain",
        headers={
            "Content-Disposition": f'attachment; filename="ejournal-{export_date.isoformat()}.txt"'
        },
    )


@readings_bp.route("/statutory-discounts/export", methods=["GET"])
@jwt_required()
def export_statutory_discounts():
    user_id = int(get_jwt_identity())
    requested_date = request.args.get("date")
    try:
        report_date = (
            datetime.strptime(requested_date, "%Y-%m-%d").date()
            if requested_date
            else business_today(User.query.get(user_id).business_day_cutoff)
        )
    except ValueError:
        return jsonify({"error": "date must use YYYY-MM-DD format"}), 400

    user = User.query.get(user_id)
    start_dt, end_dt = business_day_utc_bounds(report_date, user.business_day_cutoff)
    invoices = Invoice.query.filter(
        Invoice.user_id == user_id,
        Invoice.date_time >= start_dt,
        Invoice.date_time <= end_dt,
        Invoice.discount_amount > 0,
    ).order_by(Invoice.date_time).all()
    labels = {
        "senior_citizen": "Senior Citizen",
        "pwd": "Persons with Disability",
        "naac": "National Athletes and Coaches",
        "solo_parent": "Solo Parent",
    }
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="7F1D2D")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = [
        "Invoice No.", "Date/Time", "Beneficiary", "TIN", "ID No.",
        "Gross Sales", "Discount", "Net Sales", "Status",
    ]
    for discount_type, label in labels.items():
        ws = wb.create_sheet(label[:31])
        ws.merge_cells("A1:I1")
        ws["A1"] = user.business_name
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A2:I2")
        ws["A2"] = f"{label} Sales Book - {report_date.isoformat()}"
        ws["A2"].font = Font(bold=True)
        for column, value in enumerate(headers, start=1):
            cell = ws.cell(4, column, value)
            cell.fill = header_fill
            cell.font = white_font
            cell.border = border
        rows = [invoice for invoice in invoices if invoice.discount_type == discount_type]
        for row_index, invoice in enumerate(rows, start=5):
            values = [
                invoice.invoice_number,
                to_business_datetime(invoice.date_time).strftime("%Y-%m-%d %H:%M:%S"),
                invoice.discount_beneficiary_name or "",
                invoice.discount_beneficiary_tin or "",
                invoice.discount_id_no or "",
                float(invoice.subtotal),
                float(invoice.discount_amount),
                float(invoice.total_amount),
                invoice.status,
            ]
            for column, value in enumerate(values, start=1):
                cell = ws.cell(row_index, column, value)
                cell.border = border
                if column in (6, 7, 8):
                    cell.number_format = '"PHP" #,##0.00'
        widths = (15, 22, 28, 18, 20, 16, 16, 16, 12)
        for column, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(column)].width = width
        ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    add_audit_event(
        user_id=user_id,
        entity_type="statutory_discount_report",
        entity_id=int(report_date.strftime("%Y%m%d")),
        action="export",
        after={"format": "xlsx", "business_date": report_date},
    )
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"statutory-discount-books-{report_date.isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
